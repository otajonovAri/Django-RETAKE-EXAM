from django.test import TestCase, override_settings
from django.urls import reverse

from .models import City, Country


class CityCrudTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.uz = Country.objects.create(name="O'zbekiston", code="UZ")
        cls.jp = Country.objects.create(name="Yaponiya", code="JP")
        cls.tashkent = City.objects.create(
            name="Toshkent", population=2_900_000, country=cls.uz, is_capital=True
        )
        cls.tokyo = City.objects.create(
            name="Tokio", population=13_960_000, country=cls.jp, is_capital=True
        )

    def test_list_shows_all_cities(self):
        response = self.client.get(reverse("cities:list"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Toshkent")
        self.assertContains(response, "Tokio")
        self.assertEqual(response.context["total_cities"], 2)

    def test_filter_by_country(self):
        response = self.client.get(reverse("cities:list"), {"country": self.uz.pk})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Toshkent")
        self.assertNotContains(response, "Tokio")
        self.assertEqual(response.context["total_cities"], 1)

    def test_search_by_name(self):
        response = self.client.get(reverse("cities:list"), {"q": "tok"})
        self.assertEqual(response.context["total_cities"], 1)
        self.assertContains(response, "Tokio")

    def test_sort_by_population_desc(self):
        response = self.client.get(reverse("cities:list"), {"sort": "-population"})
        cities = list(response.context["cities"])
        self.assertEqual(cities[0], self.tokyo)

    def test_create_city(self):
        response = self.client.post(
            reverse("cities:create"),
            {
                "name": "Samarqand",
                "country": self.uz.pk,
                "population": 570_000,
                "is_capital": False,
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertTrue(City.objects.filter(name="Samarqand").exists())

    def test_create_duplicate_city_in_same_country_is_rejected(self):
        response = self.client.post(
            reverse("cities:create"),
            {
                "name": "toshkent",
                "country": self.uz.pk,
                "population": 100,
                "is_capital": False,
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(City.objects.filter(country=self.uz).count(), 1)

    def test_update_city(self):
        response = self.client.post(
            reverse("cities:update", args=[self.tashkent.pk]),
            {
                "name": "Toshkent",
                "country": self.uz.pk,
                "population": 3_000_000,
                "is_capital": True,
            },
        )
        self.assertEqual(response.status_code, 302)
        self.tashkent.refresh_from_db()
        self.assertEqual(self.tashkent.population, 3_000_000)

    def test_delete_city(self):
        response = self.client.post(reverse("cities:delete", args=[self.tokyo.pk]))
        self.assertEqual(response.status_code, 302)
        self.assertFalse(City.objects.filter(pk=self.tokyo.pk).exists())

    def test_detail_page(self):
        response = self.client.get(reverse("cities:detail", args=[self.tashkent.pk]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Toshkent")


class CountryViewTests(TestCase):
    def test_country_list_reports_counts(self):
        uz = Country.objects.create(name="O'zbekiston", code="UZ")
        City.objects.create(name="Toshkent", population=2_900_000, country=uz)
        City.objects.create(name="Samarqand", population=570_000, country=uz)

        response = self.client.get(reverse("cities:country_list"))
        self.assertEqual(response.status_code, 200)
        country = response.context["countries"][0]
        self.assertEqual(country.city_count, 2)
        self.assertEqual(country.population_sum, 3_470_000)

    def test_create_country(self):
        response = self.client.post(
            reverse("cities:country_create"), {"name": "Italiya", "code": "it"}
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(Country.objects.get(name="Italiya").code, "IT")


class CityModelTests(TestCase):
    def test_str_representation(self):
        uz = Country.objects.create(name="O'zbekiston")
        city = City.objects.create(name="Buxoro", population=280_000, country=uz)
        self.assertEqual(str(city), "Buxoro (O'zbekiston)")

    def test_deleting_country_deletes_its_cities(self):
        uz = Country.objects.create(name="O'zbekiston")
        City.objects.create(name="Xiva", population=93_000, country=uz)
        uz.delete()
        self.assertEqual(City.objects.count(), 0)


class AccessibilityTests(TestCase):
    """WCAG 2.1 AA bo'yicha tuzatilgan joylar regressiyaga uchramasligi uchun."""

    @classmethod
    def setUpTestData(cls):
        cls.uz = Country.objects.create(name="O'zbekiston", code="UZ")
        cls.city = City.objects.create(
            name="Toshkent", population=2_900_000, country=cls.uz, is_capital=True
        )

    def test_page_declares_language(self):
        html = self.client.get(reverse("cities:list")).content.decode()
        self.assertIn('<html lang="uz"', html)

    def test_skip_link_is_first_focusable_and_targets_main(self):
        html = self.client.get(reverse("cities:list")).content.decode()
        self.assertIn('class="skip-link" href="#main"', html)
        self.assertIn('id="main"', html)

    def test_nav_has_accessible_name_and_current_page(self):
        html = self.client.get(reverse("cities:list")).content.decode()
        self.assertIn('aria-label="Asosiy menyu"', html)
        self.assertIn('aria-current="page"', html)

    def test_messages_are_in_live_region(self):
        html = self.client.get(reverse("cities:list")).content.decode()
        self.assertIn('role="status"', html)
        self.assertIn('aria-live="polite"', html)

    def test_country_filter_does_not_auto_submit(self):
        """3.2.2 - tanlash o'zi sahifani yubormasligi kerak."""
        html = self.client.get(reverse("cities:list")).content.decode()
        self.assertNotIn("this.form.submit()", html)
        self.assertNotIn("onchange", html)

    def test_sortable_headers_expose_aria_sort(self):
        html = self.client.get(reverse("cities:list"), {"sort": "-population"}).content.decode()
        self.assertIn('aria-sort="descending"', html)
        self.assertIn('aria-sort="none"', html)

    def test_table_has_caption_and_row_headers(self):
        html = self.client.get(reverse("cities:list")).content.decode()
        self.assertIn("<caption>", html)
        self.assertIn('<th scope="col"', html)
        self.assertIn('<th scope="row"', html)

    def test_row_action_links_have_unique_accessible_names(self):
        html = self.client.get(reverse("cities:list")).content.decode()
        self.assertIn('<span class="visually-hidden">: Toshkent</span>', html)

    def test_invalid_field_is_marked_and_described(self):
        """3.3.1 / 4.1.2 - xato maydon aria orqali bog'lanadi."""
        response = self.client.post(
            reverse("cities:create"),
            {"name": "", "country": self.uz.pk, "population": ""},
        )
        html = response.content.decode()
        self.assertEqual(response.status_code, 200)
        self.assertIn('aria-invalid="true"', html)
        self.assertIn('aria-describedby="id_name_error"', html)
        self.assertIn('id="id_name_error"', html)
        self.assertIn('role="alert"', html)

    def test_required_fields_expose_required_attribute(self):
        html = self.client.get(reverse("cities:create")).content.decode()
        self.assertIn("required", html)
        self.assertIn("(majburiy maydon)", html)

    def test_help_text_is_associated_with_input(self):
        html = self.client.get(reverse("cities:create")).content.decode()
        self.assertIn('aria-describedby="id_population_helptext"', html)
        self.assertIn('id="id_population_helptext"', html)

    def test_population_min_matches_server_validation(self):
        html = self.client.get(reverse("cities:create")).content.decode()
        self.assertIn('min="1"', html)

    def test_pagination_nav_is_labelled(self):
        for i in range(20):
            City.objects.create(name=f"Shahar {i}", population=1000 + i, country=self.uz)
        html = self.client.get(reverse("cities:list")).content.decode()
        self.assertIn('aria-label="Sahifalar"', html)


class CountryCrudTests(TestCase):
    """Mamlakat uchun to'liq CRUD."""

    def setUp(self):
        self.uz = Country.objects.create(name="O'zbekiston", code="UZ")
        self.tashkent = City.objects.create(
            name="Toshkent", population=2_900_000, country=self.uz, is_capital=True
        )
        City.objects.create(name="Samarqand", population=570_000, country=self.uz)

    def test_detail_lists_cities_and_totals(self):
        response = self.client.get(reverse("cities:country_detail", args=[self.uz.pk]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Toshkent")
        self.assertContains(response, "Samarqand")
        self.assertEqual(response.context["city_count"], 2)
        self.assertEqual(response.context["population_sum"], 3_470_000)
        self.assertEqual(response.context["capital"], self.tashkent)

    def test_detail_orders_cities_by_population_desc(self):
        response = self.client.get(reverse("cities:country_detail", args=[self.uz.pk]))
        names = [c.name for c in response.context["cities"]]
        self.assertEqual(names, ["Toshkent", "Samarqand"])

    def test_update_country(self):
        response = self.client.post(
            reverse("cities:country_update", args=[self.uz.pk]),
            {"name": "Uzbekistan", "code": "uz"},
        )
        self.assertEqual(response.status_code, 302)
        self.uz.refresh_from_db()
        self.assertEqual(self.uz.name, "Uzbekistan")
        self.assertEqual(self.uz.code, "UZ")

    def test_delete_confirm_page_warns_about_cities(self):
        response = self.client.get(reverse("cities:country_delete", args=[self.uz.pk]))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["city_count"], 2)
        self.assertIn("Samarqand", response.context["cities_preview"])

    def test_delete_country_removes_its_cities(self):
        response = self.client.post(reverse("cities:country_delete", args=[self.uz.pk]))
        self.assertEqual(response.status_code, 302)
        self.assertEqual(Country.objects.count(), 0)
        self.assertEqual(City.objects.count(), 0)

    def test_country_list_links_to_detail(self):
        response = self.client.get(reverse("cities:country_list"))
        self.assertContains(
            response, reverse("cities:country_detail", args=[self.uz.pk])
        )

    def test_missing_country_returns_404(self):
        response = self.client.get(reverse("cities:country_detail", args=[9999]))
        self.assertEqual(response.status_code, 404)


class ExportTests(TestCase):
    """CSV va Excel eksporti joriy filtrga bo'ysunadi."""

    @classmethod
    def setUpTestData(cls):
        cls.uz = Country.objects.create(name="O'zbekiston", code="UZ")
        cls.jp = Country.objects.create(name="Yaponiya", code="JP")
        City.objects.create(
            name="Toshkent", population=2_900_000, country=cls.uz, is_capital=True
        )
        City.objects.create(name="Samarqand", population=570_000, country=cls.uz)
        City.objects.create(
            name="Tokio", population=13_960_000, country=cls.jp, is_capital=True
        )

    def _csv_rows(self, response):
        text = response.content.decode("utf-8-sig")
        return [line for line in text.splitlines() if line.strip()]

    def test_csv_headers_and_filename(self):
        response = self.client.get(reverse("cities:export_csv"))
        self.assertEqual(response.status_code, 200)
        self.assertIn("text/csv", response["Content-Type"])
        self.assertIn("attachment;", response["Content-Disposition"])
        self.assertIn(".csv", response["Content-Disposition"])

    def test_csv_starts_with_bom_for_excel(self):
        response = self.client.get(reverse("cities:export_csv"))
        self.assertTrue(response.content.startswith("\ufeff".encode("utf-8")))

    def test_csv_contains_all_cities(self):
        rows = self._csv_rows(self.client.get(reverse("cities:export_csv")))
        self.assertEqual(len(rows), 4)  # sarlavha + 3 shahar
        self.assertIn("Shahar;Mamlakat;ISO kodi;Aholisi;Poytaxt", rows[0])

    def test_csv_respects_country_filter(self):
        rows = self._csv_rows(
            self.client.get(reverse("cities:export_csv"), {"country": self.jp.pk})
        )
        self.assertEqual(len(rows), 2)  # sarlavha + Tokio
        self.assertIn("Tokio", rows[1])

    def test_csv_respects_search_filter(self):
        rows = self._csv_rows(
            self.client.get(reverse("cities:export_csv"), {"q": "samar"})
        )
        self.assertEqual(len(rows), 2)
        self.assertIn("Samarqand", rows[1])

    def test_xlsx_is_a_valid_workbook(self):
        from io import BytesIO

        from openpyxl import load_workbook

        response = self.client.get(reverse("cities:export_xlsx"))
        self.assertEqual(response.status_code, 200)
        self.assertIn("spreadsheetml.sheet", response["Content-Type"])

        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook.active
        self.assertEqual(sheet.title, "Shaharlar")
        self.assertEqual(sheet.max_row, 4)  # sarlavha + 3 shahar
        self.assertEqual(
            [c.value for c in sheet[1]],
            ["Shahar", "Mamlakat", "ISO kodi", "Aholisi", "Poytaxt"],
        )

    def test_xlsx_respects_filter(self):
        from io import BytesIO

        from openpyxl import load_workbook

        response = self.client.get(
            reverse("cities:export_xlsx"), {"country": self.uz.pk}
        )
        sheet = load_workbook(BytesIO(response.content)).active
        self.assertEqual(sheet.max_row, 3)  # sarlavha + 2 shahar

    def test_export_buttons_appear_on_list(self):
        response = self.client.get(reverse("cities:list"))
        self.assertContains(response, reverse("cities:export_csv"))
        self.assertContains(response, reverse("cities:export_xlsx"))


class ApiTests(TestCase):
    """REST API: o'qish ochiq, yozish uchun ruxsat kerak."""

    @classmethod
    def setUpTestData(cls):
        cls.uz = Country.objects.create(name="O'zbekiston", code="UZ")
        cls.jp = Country.objects.create(name="Yaponiya", code="JP")
        cls.tashkent = City.objects.create(
            name="Toshkent", population=2_900_000, country=cls.uz, is_capital=True
        )
        City.objects.create(name="Samarqand", population=570_000, country=cls.uz)
        City.objects.create(
            name="Tokio", population=13_960_000, country=cls.jp, is_capital=True
        )

    def _admin(self):
        from django.contrib.auth import get_user_model

        return get_user_model().objects.create_superuser(
            username="admin", email="admin@example.com", password="parol12345"
        )

    # --- o'qish ---

    def test_city_list_is_public(self):
        response = self.client.get("/api/cities/")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["count"], 3)

    def test_city_payload_includes_country_name(self):
        response = self.client.get(f"/api/cities/{self.tashkent.pk}/")
        data = response.json()
        self.assertEqual(data["name"], "Toshkent")
        self.assertEqual(data["country_name"], "O'zbekiston")
        self.assertEqual(data["country_code"], "UZ")

    def test_filter_by_country(self):
        response = self.client.get("/api/cities/", {"country": self.jp.pk})
        self.assertEqual(response.json()["count"], 1)

    def test_filter_by_is_capital(self):
        response = self.client.get("/api/cities/", {"is_capital": "true"})
        self.assertEqual(response.json()["count"], 2)

    def test_search(self):
        response = self.client.get("/api/cities/", {"search": "tok"})
        self.assertEqual(response.json()["count"], 1)

    def test_ordering(self):
        response = self.client.get("/api/cities/", {"ordering": "-population"})
        names = [row["name"] for row in response.json()["results"]]
        self.assertEqual(names[0], "Tokio")

    def test_country_list_includes_counts(self):
        response = self.client.get("/api/countries/")
        rows = {row["name"]: row for row in response.json()["results"]}
        self.assertEqual(rows["O'zbekiston"]["city_count"], 2)
        self.assertEqual(rows["O'zbekiston"]["population_sum"], 3_470_000)

    def test_country_cities_action(self):
        response = self.client.get(f"/api/countries/{self.uz.pk}/cities/")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["count"], 2)

    # --- yozish ---

    def test_anonymous_cannot_create(self):
        response = self.client.post(
            "/api/cities/",
            {"name": "Buxoro", "population": 280000, "country": self.uz.pk},
            content_type="application/json",
        )
        self.assertIn(response.status_code, (401, 403))
        self.assertFalse(City.objects.filter(name="Buxoro").exists())

    def test_anonymous_cannot_delete(self):
        response = self.client.delete(f"/api/cities/{self.tashkent.pk}/")
        self.assertIn(response.status_code, (401, 403))
        self.assertTrue(City.objects.filter(pk=self.tashkent.pk).exists())

    def test_admin_can_create(self):
        self.client.force_login(self._admin())
        response = self.client.post(
            "/api/cities/",
            {"name": "Buxoro", "population": 280000, "country": self.uz.pk},
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 201, response.content)
        self.assertTrue(City.objects.filter(name="Buxoro").exists())

    def test_admin_can_update_and_delete(self):
        self.client.force_login(self._admin())

        patch = self.client.patch(
            f"/api/cities/{self.tashkent.pk}/",
            {"population": 3_000_000},
            content_type="application/json",
        )
        self.assertEqual(patch.status_code, 200)
        self.tashkent.refresh_from_db()
        self.assertEqual(self.tashkent.population, 3_000_000)

        delete = self.client.delete(f"/api/cities/{self.tashkent.pk}/")
        self.assertEqual(delete.status_code, 204)
        self.assertFalse(City.objects.filter(pk=self.tashkent.pk).exists())

    def test_api_rejects_duplicate_city_in_same_country(self):
        self.client.force_login(self._admin())
        response = self.client.post(
            "/api/cities/",
            {"name": "toshkent", "population": 100, "country": self.uz.pk},
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn("name", response.json())

    def test_api_rejects_zero_population(self):
        self.client.force_login(self._admin())
        response = self.client.post(
            "/api/cities/",
            {"name": "Nol", "population": 0, "country": self.uz.pk},
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 400)

    # --- hujjatlar ---

    def test_openapi_schema_is_served(self):
        response = self.client.get("/api/schema/")
        self.assertEqual(response.status_code, 200)

    def test_swagger_and_redoc_pages(self):
        self.assertEqual(self.client.get("/api/docs/").status_code, 200)
        self.assertEqual(self.client.get("/api/redoc/").status_code, 200)

    def test_swagger_assets_are_served_locally(self):
        """Internetsiz muhitda ham ochilishi uchun CDN ishlatilmasin."""
        html = self.client.get("/api/docs/").content.decode()
        self.assertIn("/static/drf_spectacular_sidecar/", html)
        self.assertNotIn("cdn.jsdelivr.net", html)
        self.assertNotIn("unpkg.com", html)

    def test_schema_lists_expected_endpoints(self):
        import json

        schema = json.loads(self.client.get("/api/schema/?format=json").content)
        paths = set(schema["paths"])
        for expected in (
            "/api/cities/",
            "/api/cities/{id}/",
            "/api/countries/",
            "/api/countries/{id}/",
            "/api/countries/{id}/cities/",
        ):
            self.assertIn(expected, paths)


@override_settings(DEBUG=False)
class ErrorPageTests(TestCase):
    """Maxsus xato sahifalari DEBUG=False bo'lganda ishlaydi."""

    def test_custom_404_page(self):
        response = self.client.get("/bunday-sahifa-yoq/")
        self.assertEqual(response.status_code, 404)
        self.assertContains(response, "Sahifa topilmadi", status_code=404)
        self.assertContains(response, "Bosh sahifaga qaytish", status_code=404)

    def test_404_page_uses_site_layout(self):
        response = self.client.get("/bunday-sahifa-yoq/")
        self.assertContains(response, "Dunyo shaharlari", status_code=404)


@override_settings(DEBUG=True)
class CreateAdminCommandTests(TestCase):
    """`createadmin` buyrug'i — admin foydalanuvchi yaratish.

    Django test paytida DEBUG=False qiladi, buyruq esa bunday holatda
    standart parolni rad etadi. Shuning uchun ishlab chiqish holatini
    ataylab qaytaramiz; production xatti-harakati alohida tekshiriladi.
    """

    def setUp(self):
        from django.contrib.auth import get_user_model

        self.User = get_user_model()

    def _run(self, **kwargs):
        from io import StringIO

        from django.core.management import call_command

        out, err = StringIO(), StringIO()
        call_command("createadmin", stdout=out, stderr=err, **kwargs)
        return out.getvalue(), err.getvalue()

    def test_creates_superuser_with_defaults(self):
        out, _ = self._run()
        user = self.User.objects.get(username="admin")
        self.assertTrue(user.is_superuser)
        self.assertTrue(user.is_staff)
        self.assertTrue(user.check_password("admin12345"))
        self.assertIn("Admin yaratildi", out)

    def test_is_idempotent(self):
        self._run()
        out, _ = self._run()
        self.assertEqual(self.User.objects.filter(username="admin").count(), 1)
        self.assertIn("allaqachon mavjud", out)

    def test_reset_password(self):
        self._run()
        self._run(password="yangiParol123", reset_password=True)
        user = self.User.objects.get(username="admin")
        self.assertTrue(user.check_password("yangiParol123"))

    def test_custom_username_and_password(self):
        self._run(username="boshqa", password="Parol!2345", email="a@b.uz")
        user = self.User.objects.get(username="boshqa")
        self.assertTrue(user.check_password("Parol!2345"))
        self.assertEqual(user.email, "a@b.uz")

    @override_settings(DEBUG=False)
    def test_refuses_default_password_in_production(self):
        _, err = self._run()
        self.assertIn("standart parol ishlatilmaydi", err)
        self.assertFalse(self.User.objects.filter(username="admin").exists())

    @override_settings(DEBUG=False)
    def test_allows_explicit_password_in_production(self):
        self._run(password="JudaKuchliParol!2026")
        self.assertTrue(self.User.objects.filter(username="admin").exists())


class AdminSiteTests(TestCase):
    """Admin panel va uning modellari."""

    def setUp(self):
        from django.contrib.auth import get_user_model

        self.user = get_user_model().objects.create_superuser(
            username="admin", email="admin@example.com", password="admin12345"
        )
        self.uz = Country.objects.create(name="O'zbekiston", code="UZ")
        City.objects.create(name="Toshkent", population=2_900_000, country=self.uz)

    def test_login_page_is_reachable(self):
        self.assertEqual(self.client.get("/admin/login/").status_code, 200)

    def test_admin_requires_login(self):
        response = self.client.get("/admin/cities/city/")
        self.assertEqual(response.status_code, 302)
        self.assertIn("/admin/login/", response["Location"])

    def test_admin_index_lists_both_models(self):
        self.client.force_login(self.user)
        response = self.client.get("/admin/")
        self.assertContains(response, "/admin/cities/city/")
        self.assertContains(response, "/admin/cities/country/")

    def test_city_changelist_and_search(self):
        self.client.force_login(self.user)
        response = self.client.get("/admin/cities/city/", {"q": "Toshkent"})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Toshkent")

    def test_country_changelist(self):
        self.client.force_login(self.user)
        response = self.client.get("/admin/cities/country/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "O&#x27;zbekiston")


class ApiAuthSchemeTests(TestCase):
    """Swagger'dagi «Authorize» ishlashi uchun sxemada auth e'lon qilinsin."""

    def test_schema_declares_basic_and_cookie_auth(self):
        import json

        schema = json.loads(self.client.get("/api/schema/?format=json").content)
        schemes = schema["components"]["securitySchemes"]
        self.assertIn("basicAuth", schemes)
        self.assertIn("cookieAuth", schemes)

    def test_basic_auth_allows_write(self):
        import base64

        from django.contrib.auth import get_user_model

        get_user_model().objects.create_superuser(
            username="admin", email="a@b.uz", password="admin12345"
        )
        country = Country.objects.create(name="Italiya", code="IT")

        token = base64.b64encode(b"admin:admin12345").decode()
        response = self.client.post(
            "/api/cities/",
            {"name": "Rim", "population": 2_800_000, "country": country.pk},
            content_type="application/json",
            HTTP_AUTHORIZATION=f"Basic {token}",
        )
        self.assertEqual(response.status_code, 201, response.content)
        self.assertTrue(City.objects.filter(name="Rim").exists())

    def test_wrong_password_is_rejected(self):
        import base64

        from django.contrib.auth import get_user_model

        get_user_model().objects.create_superuser(
            username="admin", email="a@b.uz", password="admin12345"
        )
        country = Country.objects.create(name="Italiya", code="IT")

        token = base64.b64encode(b"admin:notogri").decode()
        response = self.client.post(
            "/api/cities/",
            {"name": "Rim", "population": 100, "country": country.pk},
            content_type="application/json",
            HTTP_AUTHORIZATION=f"Basic {token}",
        )
        self.assertIn(response.status_code, (401, 403))
        self.assertFalse(City.objects.filter(name="Rim").exists())
