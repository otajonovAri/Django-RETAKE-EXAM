import csv
from datetime import date

from django.contrib import messages
from django.db.models import Count, Q, Sum
from django.http import HttpResponse
from django.shortcuts import render
from django.urls import reverse, reverse_lazy
from django.views.generic import (
    CreateView,
    DeleteView,
    DetailView,
    ListView,
    UpdateView,
)

from .forms import CityForm, CountryForm
from .models import City, Country

SORT_FIELDS = {
    "name": "name",
    "-name": "-name",
    "population": "population",
    "-population": "-population",
    "country": "country__name",
    "-country": "-country__name",
}

# Jadval sarlavhalari: (kalit, ko'rinadigan nom, raqamli ustunmi)
SORTABLE_COLUMNS = (
    ("name", "Shahar nomi", False),
    ("country", "Mamlakat", False),
    ("population", "Aholisi", True),
)

EXPORT_HEADERS = ["Shahar", "Mamlakat", "ISO kodi", "Aholisi", "Poytaxt"]


def filter_cities(params):
    """Ro'yxat va eksport bir xil filtrdan foydalanishi uchun umumiy funksiya.

    `params` — `request.GET` (QueryDict). Mamlakat, qidiruv va saralash
    parametrlari qo'llanadi.
    """
    queryset = City.objects.select_related("country")

    country_id = params.get("country")
    if country_id and country_id.isdigit():
        queryset = queryset.filter(country_id=int(country_id))

    query = params.get("q", "").strip()
    if query:
        queryset = queryset.filter(
            Q(name__icontains=query) | Q(country__name__icontains=query)
        )

    sort = params.get("sort", "")
    if sort in SORT_FIELDS:
        queryset = queryset.order_by(SORT_FIELDS[sort])

    return queryset


def city_export_rows(queryset):
    """Eksport uchun qatorlar generatori."""
    for city in queryset:
        yield [
            city.name,
            city.country.name,
            city.country.code or "",
            city.population,
            "Ha" if city.is_capital else "Yo'q",
        ]


def _export_filename(extension):
    return f"shaharlar-{date.today():%Y-%m-%d}.{extension}"


class CityListView(ListView):
    """Shaharlar ro'yxati - jadval ko'rinishida, mamlakat bo'yicha filtrlanadi."""

    model = City
    template_name = "cities/city_list.html"
    context_object_name = "cities"
    paginate_by = 15

    def get_queryset(self):
        return filter_cities(self.request.GET)

    def _sort_headers(self, sort, base_params):
        """Har bir ustun uchun aria-sort holati va keyingi saralash havolasi."""
        headers = []
        for key, label, numeric in SORTABLE_COLUMNS:
            ascending = sort == key
            descending = sort == f"-{key}"

            if numeric:
                next_sort = key if descending else f"-{key}"
            else:
                next_sort = f"-{key}" if ascending else key

            params = base_params.copy()
            params["sort"] = next_sort

            direction = "kamayish" if next_sort.startswith("-") else "o'sish"

            headers.append({
                "key": key,
                "label": label,
                "numeric": numeric,
                "aria_sort": (
                    "ascending" if ascending
                    else "descending" if descending
                    else "none"
                ),
                "url": f"?{params.urlencode()}",
                "action": f"{label} bo'yicha {direction} tartibida saralash",
            })
        return headers

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        filtered = self.get_queryset()

        selected_country = self.request.GET.get("country", "")
        query = self.request.GET.get("q", "")
        sort = self.request.GET.get("sort", "")

        # Sahifalash va saralash havolalarida filtr yo'qolmasligi uchun.
        base_params = self.request.GET.copy()
        base_params.pop("page", None)
        base_params.pop("sort", None)

        page_params = self.request.GET.copy()
        page_params.pop("page", None)

        context.update({
            "nav_section": "cities",
            "countries": Country.objects.annotate(
                city_count=Count("cities")
            ).order_by("name"),
            "selected_country": selected_country,
            "query": query,
            "sort": sort,
            "total_cities": filtered.count(),
            "total_population": filtered.aggregate(total=Sum("population"))["total"] or 0,
            "is_filtered": bool(selected_country or query),
            "querystring": page_params.urlencode(),
            "export_querystring": page_params.urlencode(),
            "sort_headers": self._sort_headers(sort, base_params),
        })
        return context


class CityDetailView(DetailView):
    model = City
    template_name = "cities/city_detail.html"
    context_object_name = "city"
    queryset = City.objects.select_related("country")
    extra_context = {"nav_section": "cities"}


class CityCreateView(CreateView):
    model = City
    form_class = CityForm
    template_name = "cities/city_form.html"
    extra_context = {
        "title": "Yangi shahar qo'shish",
        "submit_label": "Qo'shish",
        "nav_section": "cities",
    }

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, f"«{self.object.name}» shahri qo'shildi.")
        return response


class CityUpdateView(UpdateView):
    model = City
    form_class = CityForm
    template_name = "cities/city_form.html"
    extra_context = {
        "title": "Shaharni tahrirlash",
        "submit_label": "Saqlash",
        "nav_section": "cities",
    }

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, f"«{self.object.name}» ma'lumotlari yangilandi.")
        return response


class CityDeleteView(DeleteView):
    model = City
    template_name = "cities/city_confirm_delete.html"
    success_url = reverse_lazy("cities:list")
    context_object_name = "city"
    extra_context = {"nav_section": "cities"}

    def form_valid(self, form):
        messages.success(self.request, f"«{self.object.name}» shahri o'chirildi.")
        return super().form_valid(form)


class CountryListView(ListView):
    """Mamlakatlar kesimidagi statistika."""

    model = Country
    template_name = "cities/country_list.html"
    context_object_name = "countries"
    extra_context = {"nav_section": "countries"}

    def get_queryset(self):
        return Country.objects.annotate(
            city_count=Count("cities"),
            population_sum=Sum("cities__population"),
        ).order_by("-city_count", "name")


class CountryDetailView(DetailView):
    """Bitta mamlakat va uning shaharlari."""

    model = Country
    template_name = "cities/country_detail.html"
    context_object_name = "country"
    extra_context = {"nav_section": "countries"}

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        cities = self.object.cities.order_by("-population")
        context["cities"] = cities
        context["city_count"] = cities.count()
        context["population_sum"] = (
            cities.aggregate(total=Sum("population"))["total"] or 0
        )
        context["capital"] = cities.filter(is_capital=True).first()
        return context


class CountryCreateView(CreateView):
    model = Country
    form_class = CountryForm
    template_name = "cities/country_form.html"
    success_url = reverse_lazy("cities:country_list")
    extra_context = {
        "title": "Yangi mamlakat qo'shish",
        "submit_label": "Qo'shish",
        "nav_section": "countries",
    }

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, f"«{self.object.name}» mamlakati qo'shildi.")
        return response


class CountryUpdateView(UpdateView):
    model = Country
    form_class = CountryForm
    template_name = "cities/country_form.html"
    extra_context = {
        "title": "Mamlakatni tahrirlash",
        "submit_label": "Saqlash",
        "nav_section": "countries",
    }

    def get_success_url(self):
        return reverse("cities:country_detail", args=[self.object.pk])

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, f"«{self.object.name}» yangilandi.")
        return response


class CountryDeleteView(DeleteView):
    """Mamlakatni o'chirish.

    Mamlakat o'chirilsa, unga bog'liq barcha shaharlar ham o'chadi (CASCADE),
    shuning uchun tasdiqlash sahifasida nechta shahar yo'qolishi aytiladi.
    """

    model = Country
    template_name = "cities/country_confirm_delete.html"
    success_url = reverse_lazy("cities:country_list")
    context_object_name = "country"
    extra_context = {"nav_section": "countries"}

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        names = list(self.object.cities.order_by("name").values_list("name", flat=True))
        context["city_count"] = len(names)
        preview = ", ".join(names[:8])
        if len(names) > 8:
            preview += f" va yana {len(names) - 8} ta"
        context["cities_preview"] = preview
        return context

    def form_valid(self, form):
        city_count = self.object.cities.count()
        name = self.object.name
        response = super().form_valid(form)
        if city_count:
            messages.success(
                self.request,
                f"«{name}» va unga tegishli {city_count} ta shahar o'chirildi.",
            )
        else:
            messages.success(self.request, f"«{name}» mamlakati o'chirildi.")
        return response


# ------------------------------------------------------------------ eksport


def export_cities_csv(request):
    """Filtrlangan shaharlarni CSV holida yuklab beradi (Excel uchun BOM bilan)."""
    queryset = filter_cities(request.GET)

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = (
        f'attachment; filename="{_export_filename("csv")}"'
    )
    # BOM - Excel UTF-8 ni to'g'ri tanishi uchun.
    response.write("\ufeff")

    writer = csv.writer(response, delimiter=";")
    writer.writerow(EXPORT_HEADERS)
    for row in city_export_rows(queryset):
        writer.writerow(row)

    return response


def export_cities_xlsx(request):
    """Filtrlangan shaharlarni .xlsx holida yuklab beradi."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    queryset = filter_cities(request.GET)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Shaharlar"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="4338CA")

    sheet.append(EXPORT_HEADERS)
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row in city_export_rows(queryset):
        sheet.append(row)

    # Aholi ustunini raqam formatida ko'rsatamiz.
    for row in sheet.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            cell.number_format = "#,##0"

    widths = [22, 22, 10, 14, 10]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    sheet.freeze_panes = "A2"

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    )
    response["Content-Disposition"] = (
        f'attachment; filename="{_export_filename("xlsx")}"'
    )
    workbook.save(response)
    return response


# ------------------------------------------------------------ xato sahifalari


def error_404(request, exception=None):
    return render(request, "errors/404.html", status=404)


def error_403(request, exception=None):
    return render(request, "errors/403.html", status=403)


def error_500(request):
    return render(request, "errors/500.html", status=500)
